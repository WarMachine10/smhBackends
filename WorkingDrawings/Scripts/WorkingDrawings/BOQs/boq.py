# %%
import os
from django.conf import settings
import pandas as pd
import numpy as np
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import ezdxf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils.dataframe import dataframe_to_rows

# %%
def count_columns(column_info_df: dict) -> int:
    """
    Function to count the number of columns in a DataFrame.

    Parameters:
    column_info_df (dict): The input dictionary to be converted into a DataFrame.

    Returns:
    int: The number of columns in the DataFrame.
    """
    # Convert the dictionary into a DataFrame
    df = pd.DataFrame(column_info_df)
    
    # Count the number of columns
    num_columns = len(df.columns)
    return num_columns

# %%
def generate_material_quantities_col(num_columns: int, excel_file: str) -> dict:
    """
    Function to calculate the material quantities required for the given number of columns.
    
    Parameters:
    num_columns (int): The number of columns for which material quantities are to be calculated.
    excel_file (str): The path to the Excel file containing material data.
    
    Returns:
    dict: A dictionary containing the Material ID as keys and their corresponding quantities in kg as values.
    """
    # Load material details from the provided Excel file
    material_data = pd.read_excel(excel_file)

    # Steel bar weights per meter (in kg)
    steel_weights_per_meter = {
        10: 0.617,   # 10mm diameter bar weight per meter in kg
        12: 0.89,    # 12mm diameter bar weight per meter in kg
        16: 1.58     # 16mm diameter bar weight per meter in kg
    }

    # Concrete mix ratio for M20 (1:1.5:3)
    density_cement = 1440    # kg/m³
    density_sand = 1600      # kg/m³
    density_aggregate = 1450 # kg/m³
    cement_ratio = 1
    sand_ratio = 1.5
    aggregate_ratio = 3
    total_ratio = cement_ratio + sand_ratio + aggregate_ratio

    # Conversion factor from inches to meters
    INCH_TO_METER = 0.0254

    # Column dimensions in inches
    column_height_inch = 120
    column_length_inch = 12
    column_width_inch = 9

    # Convert column dimensions from inches to meters
    column_height = column_height_inch * INCH_TO_METER
    column_length = column_length_inch * INCH_TO_METER
    column_width = column_width_inch * INCH_TO_METER

    # Function to calculate the weight of steel bars
    def calculate_steel_weight(diameter_mm, length_m, num_bars):
        # Retrieve the weight per meter for the given diameter
        weight_per_meter = steel_weights_per_meter.get(diameter_mm, 0)
        total_weight_kg = weight_per_meter * length_m * num_bars  # Total weight for all bars
        return total_weight_kg

    # Function to calculate the volume of steel based on weight
    def calculate_steel_volume(steel_weight_kg):
        density_steel = 7850  # kg/m³ (Density of steel)
        steel_volume_m3 = steel_weight_kg / density_steel  # Volume in cubic meters
        return steel_volume_m3

    # Function to calculate the volume of concrete materials (cement, sand, aggregate)
    def calculate_concrete_materials(column_volume):
        concrete_volume = column_volume  # Total concrete volume (steel volume not subtracted here)
        
        cement_volume = (cement_ratio / total_ratio) * concrete_volume
        sand_volume = (sand_ratio / total_ratio) * concrete_volume
        aggregate_volume = (aggregate_ratio / total_ratio) * concrete_volume
        
        cement_weight = cement_volume * density_cement  # Cement weight in kg
        sand_weight = sand_volume * density_sand        # Sand weight in kg
        aggregate_weight = aggregate_volume * density_aggregate  # Aggregate weight in kg
        
        return cement_weight, sand_weight, aggregate_weight

    # Calculate total steel and concrete requirements for all columns
    total_steel_weight_16mm = 0
    total_steel_weight_12mm = 0
    total_cement_weight_all_columns = 0
    total_sand_weight_all_columns = 0
    total_aggregate_weight_all_columns = 0

    for _ in range(num_columns):
        # Volume of one column (m³)
        column_volume = column_height * column_length * column_width

        # Steel weight calculations for 16mm and 12mm bars
        weight_16mm_steel = calculate_steel_weight(16, column_height, 2)  # 2 bars of 16mm per column
        weight_12mm_steel = calculate_steel_weight(12, column_height, 4)  # 4 bars of 12mm per column
        
        total_steel_weight_16mm += weight_16mm_steel  # Accumulate total weight of 16mm steel in kg
        total_steel_weight_12mm += weight_12mm_steel  # Accumulate total weight of 12mm steel in kg

        # Calculate steel volume
        steel_volume_16mm = calculate_steel_volume(weight_16mm_steel)
        steel_volume_12mm = calculate_steel_volume(weight_12mm_steel)

        # Adjusted concrete volume by subtracting steel volume
        adjusted_column_volume = column_volume - (steel_volume_16mm + steel_volume_12mm)

        # Concrete material calculations using the adjusted column volume
        cement_weight, sand_weight, aggregate_weight = calculate_concrete_materials(adjusted_column_volume)

        # Accumulate the materials required for all columns
        total_cement_weight_all_columns += cement_weight
        total_sand_weight_all_columns += sand_weight
        total_aggregate_weight_all_columns += aggregate_weight

    # Prepare the material quantities dictionary
    material_quantities = {}

    # Matching the material names to their respective IDs and quantities
    for material_name, weight in {
        'Cement': total_cement_weight_all_columns,
        'Sand': total_sand_weight_all_columns,
        'Coarse Aggregate': total_aggregate_weight_all_columns
    }.items():
        matched_rows = material_data.loc[material_data['Material Name'].str.contains(material_name, case=False, na=False)]
        if not matched_rows.empty:
            material_id = matched_rows['Material ID'].values[0]
            material_quantities[material_id] = round(weight, 2)  # Round to 2 decimal places
        else:
            print(f"Material '{material_name}' not found in the Excel sheet.")

    # Matching steel bars to material IDs
    steel_bars_16mm = material_data.loc[material_data['Material Name'].str.contains("16mm", case=False, na=False)]
    steel_bars_12mm = material_data.loc[material_data['Material Name'].str.contains("12mm", case=False, na=False)]

    if not steel_bars_16mm.empty:
        steel_bars_16mm_id = steel_bars_16mm['Material ID'].values[0]
        material_quantities[steel_bars_16mm_id] = round(total_steel_weight_16mm, 2)  # Round to 2 decimal places
    else:
        print("16mm Steel Bar not found in the material list.")

    if not steel_bars_12mm.empty:
        steel_bars_12mm_id = steel_bars_12mm['Material ID'].values[0]
        material_quantities[steel_bars_12mm_id] = round(total_steel_weight_12mm, 2)  # Round to 2 decimal places
    else:
        print("12mm Steel Bar not found in the material list.")

    return material_quantities


# %%
def calculate_total_length(beam_info_df: dict) -> float:
    """
    Function to calculate the total length from the 'length' column of a dataframe.

    Parameters:
    column_info_df (dict): Dictionary containing a 'length' column with float values.

    Returns:
    float: Total sum of the 'length' column as a float.
    """
    # Convert the dictionary into a DataFrame
    df = pd.DataFrame(beam_info_df)
    
    # Sum all values in the 'length' column and return as a float
    total_length = df['length'].sum()
    return float(total_length)


# %%
def calculate_material_quantities_beam(total_beam_length_inch: float, excel_file: str) -> dict:
    """
    Calculate the material quantities (steel and concrete) required for a beam based on the beam length.

    Parameters:
    total_beam_length_inch (float): Length of the beam in inches.
    excel_file (str): Path to the Excel file containing material data.

    Returns:
    dict: A dictionary with material IDs as keys and quantities (in kg) as values, rounded to 2 decimal places.
    """
    # Steel bar information (diameter in mm and quantity per type)
    steel_bars = [
        {"diameter": 10, "quantity": 2, "material_id": "M-004"},  # Steel Bars 10mm
        {"diameter": 12, "quantity": 2, "material_id": "M-003"},  # Steel Bars 12mm
        {"diameter": 16, "quantity": 2, "material_id": "M-005"}   # Steel Bars 16mm
    ]

    # Beam dimensions (constant width and depth in mm)
    beam_width = 200  # mm
    beam_depth = 450  # mm

    # Convert inches to meters for calculations (1 inch = 0.0254 meters)
    beam_length = total_beam_length_inch * 0.0254

    # Constants for material densities (in kg/m³)
    density_cement = 1440  # kg/m³
    density_sand = 1600    # kg/m³
    density_aggregate = 1450  # kg/m³

    # Steel bar weight per meter (in kg)
    steel_bar_weights = {
        10: 0.617,   # 10mm diameter bar weight per meter in kg
        12: 0.89,    # 12mm diameter bar weight per meter in kg
        16: 1.58     # 16mm diameter bar weight per meter in kg
    }

    # Calculate the total weight of steel required (weight per piece for steel bars)
    total_steel_weight = 0
    total_steel_volume = 0

    # Calculate the total volume of steel (in m³) and weight for each steel bar
    for bar in steel_bars:
        diameter = bar["diameter"]
        quantity = bar["quantity"]
        
        # Convert diameter to meters
        diameter_m = diameter / 1000
        
        # Calculate volume per bar using the formula V = π * (d/2)² * L
        volume_per_bar = math.pi * (diameter_m / 2) ** 2 * beam_length  # volume in m³
        
        # Total volume for the given quantity of this diameter
        total_steel_volume += volume_per_bar * quantity
        
        # Calculate weight per bar using the weight per meter
        weight_per_bar = steel_bar_weights[diameter] * beam_length
        
        # Total weight for the given quantity of this diameter
        total_steel_weight += weight_per_bar * quantity

    # Calculate the total volume of the beam (in m³)
    beam_width_m = beam_width / 1000  # Convert mm to m
    beam_depth_m = beam_depth / 1000  # Convert mm to m

    beam_volume = beam_width_m * beam_depth_m * beam_length  # Volume in m³

    # Subtract the steel volume from the beam volume to get the concrete volume
    concrete_volume = beam_volume - total_steel_volume

    # Calculate the quantities of cement, sand, and aggregate for M20 grade concrete (1:1.5:3)
    cement_ratio = 1
    sand_ratio = 1.5
    aggregate_ratio = 3
    total_ratio = cement_ratio + sand_ratio + aggregate_ratio

    # Concrete material volumes based on the ratio
    cement_volume = (cement_ratio / total_ratio) * concrete_volume
    sand_volume = (sand_ratio / total_ratio) * concrete_volume
    aggregate_volume = (aggregate_ratio / total_ratio) * concrete_volume

    # Convert volumes to masses (in kg)
    cement_mass = cement_volume * density_cement
    sand_mass = sand_volume * density_sand
    aggregate_mass = aggregate_volume * density_aggregate

    # Read material data from the Excel file
    material_data = pd.read_excel(excel_file)

    # Create a dictionary to store Material IDs and quantities
    material_dict = {}

    # Add steel bars to the dictionary (per piece)
    for bar in steel_bars:
        material_dict[bar["material_id"]] = round(bar["quantity"] * steel_bar_weights[bar["diameter"]] * beam_length, 2)  # Quantity in kg with 2 decimal places

    # Add concrete materials to the dictionary
    for material_name, material_mass in zip(
            ['Cement', 'Sand', 'Coarse Aggregate'],  # Matching with your required materials
            [cement_mass, sand_mass, aggregate_mass]):

        # Find the row in the Excel sheet that matches the material name
        material_row = material_data[material_data['Material Name'].str.contains(material_name, case=False)].iloc[0]
        
        # Get Material ID from the matched row
        material_id = material_row['Material ID']
        
        # Store the Material ID and Quantity in the dictionary (in kg with 2 decimal places)
        material_dict[material_id] = round(material_mass, 2)

    return material_dict


# Get the material quantities


# %%
def concat(material_quantities_col, material_quantities_beam):
    # Convert dictionaries to DataFrames, ensuring 'Material ID' and 'Quantity' columns
    df_col = pd.DataFrame(list(material_quantities_col.items()), columns=['Material ID', 'Quantity_col'])
    df_beam = pd.DataFrame(list(material_quantities_beam.items()), columns=['Material ID', 'Quantity_beam'])

    # Merge the dataframes on 'Material ID' and sum the quantities for the same Material ID
    df_merged = pd.merge(df_col, df_beam, on='Material ID', how='outer')

    # Fill NaN values with 0 (for missing data)
    df_merged.fillna(0, inplace=True)

    # Sum the quantities for the same Material ID
    df_merged['Total Quantity'] = df_merged['Quantity_col'] + df_merged['Quantity_beam']

    # Final dictionary with summed quantities
    material_quantities_final = dict(zip(df_merged['Material ID'], df_merged['Total Quantity']))

    return material_quantities_final

# %%
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Function to load data based on the phase and quantities dictionary
def load_data(file_path, quantities):
    data = pd.read_excel(file_path)
    data['Quantity'] = data['Material ID'].map(quantities).fillna(0)
    data['Total Price (INR)'] = data['Quantity'] * data['Unit Price (INR)']
    return data

# Function to add BOQ data to the worksheet
def add_boq_data(sheet, data, title, start_row=5):
    sheet.merge_cells(f"A{start_row-1}:H{start_row-1}")
    sheet[f"A{start_row-1}"] = title
    sheet[f"A{start_row-1}"].font = Font(bold=True, size=12)
    sheet[f"A{start_row-1}"].alignment = Alignment(horizontal="center")

    # Adding rows of data
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            # Adjust column width for readability
            sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Adjust column width for 'Material Name' (second column)
            if c_idx == 2:  # Column B (Material Name)
                sheet.column_dimensions['B'].width = 30  # Increase width of the Material Name column
            else:
                sheet.column_dimensions[chr(64 + c_idx)].width = 15  # Set column width for other columns

            # Align all cells to the center
            sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")

    # Formatting header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FFA500")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border_style = Side(border_style="thin", color="000000")

    # Header formatting
    for cell in sheet[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Formatting all cells in the table
    for row in sheet.iter_rows(min_row=start_row+1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            # Align all cells to the center
            cell.alignment = Alignment(horizontal="center", vertical="center")

    total_amount = data['Total Price (INR)'].sum()
    return total_amount

# Function to generate and save the Structural BOQ
def generate_structural_boq(quantities):
    wb = Workbook()
    # Remove the default sheet created and add a new one for Structural Phase
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Structural Phase")
    xlPath=os.path.join(settings.BASE_DIR,'assets','price_of_material_new.xlsx')
    # Load and process the structural data
    structural_data = load_data(xlPath, quantities)
    total_structural = add_boq_data(ws, structural_data, "Structural Phase")
    
    # Add the Grand Total row
    ws.merge_cells(f"B{ws.max_row + 1}:E{ws.max_row + 1}")
    ws[f"B{ws.max_row}"] = "Phase Grand Total"
    ws[f"B{ws.max_row}"].font = Font(bold=True)
    ws[f"B{ws.max_row}"].alignment = Alignment(horizontal="right")
    ws[f"F{ws.max_row}"] = total_structural
    ws[f"F{ws.max_row}"].font = Font(bold=True)

    # Save the workbook with no empty sheet
    wb.save("Structural_BOQ.xlsx")



# %%
def main_structure(column_info_df, beam_info_df, excel_file):
    num_columns = count_columns(column_info_df)
    material_quantities_col = generate_material_quantities_col(num_columns, excel_file)
    total_beam_length_inch = calculate_total_length(beam_info_df)
    material_quantities_beam = calculate_material_quantities_beam(total_beam_length_inch, excel_file)
    structural_quantities = concat(material_quantities_col, material_quantities_beam)
    generate_structural_boq(structural_quantities)
    return "Structural BOQ generated successfully and saved as 'Structural_BOQ.xlsx'"

# Function to load data based on the phase and quantities dictionary
def load_data(file_path, quantities):
    data = pd.read_excel(file_path)
    data['Quantity'] = data['Material ID'].map(quantities).fillna(0)
    data['Total Price (INR)'] = data['Quantity'] * data['Unit Price (INR)']
    return data

# Function to add BOQ data to the worksheet
def add_boq_data(sheet, data, title, start_row=5):
    sheet.merge_cells(f"A{start_row-1}:H{start_row-1}")
    sheet[f"A{start_row-1}"] = title
    sheet[f"A{start_row-1}"].font = Font(bold=True, size=12)
    sheet[f"A{start_row-1}"].alignment = Alignment(horizontal="center")

    # Adding rows of data
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            # Adjust column width for readability
            sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Adjust column width for 'Material Name' (second column)
            if c_idx == 2:  # Column B (Material Name)
                sheet.column_dimensions['B'].width = 30  # Increase width of the Material Name column
            else:
                sheet.column_dimensions[chr(64 + c_idx)].width = 15  # Set column width for other columns

            # Align all cells to the center
            sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")

    # Formatting header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FFA500")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border_style = Side(border_style="thin", color="000000")

    # Header formatting
    for cell in sheet[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Formatting all cells in the table
    for row in sheet.iter_rows(min_row=start_row+1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            # Align all cells to the center
            cell.alignment = Alignment(horizontal="center", vertical="center")

    total_amount = data['Total Price (INR)'].sum()
    return total_amount

# Function to generate and save the Plumbing BOQ
def generate_plumbing_boq(quantities):
    wb = Workbook()
    # Remove the default sheet created and add a new one for Plumbing BOQ
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Plumbing BOQ")
    
    # Load and process the plumbing data
    xlPath=os.path.join(settings.BASE_DIR,'assets','price_of_plumbing_material.xlsx')
    plumbing_data = load_data(xlPath, quantities)
    total_plumbing = add_boq_data(ws, plumbing_data, "PlumbingBOQ")
    
    # Add the Grand Total row
    ws.merge_cells(f"B{ws.max_row + 1}:E{ws.max_row + 1}")
    ws[f"B{ws.max_row}"] = "Grand Total"
    ws[f"B{ws.max_row}"].font = Font(bold=True)
    ws[f"B{ws.max_row}"].alignment = Alignment(horizontal="right")
    ws[f"F{ws.max_row}"] = total_plumbing
    ws[f"F{ws.max_row}"].font = Font(bold=True)

    # Save the workbook with no empty sheet
    wb.save("Plumbing_BOQ.xlsx")

# %%
def main_plumbing(MaterialList):
    generate_plumbing_boq(MaterialList)
    # Return a success message
    return "Plumbing BOQ generated successfully and saved as 'Plumbing_BOQ.xlsx'"

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Function to load data based on the phase and quantities dictionary
def load_data(file_path, quantities):
    data = pd.read_excel(file_path)
    data['Quantity'] = data['Material ID'].map(quantities).fillna(0)
    data['Total Price (INR)'] = data['Quantity'] * data['Unit Price (INR)']
    return data

# Function to add BOQ data to the worksheet
def add_boq_data(sheet, data, title, start_row=5):
    sheet.merge_cells(f"A{start_row-1}:H{start_row-1}")
    sheet[f"A{start_row-1}"] = title
    sheet[f"A{start_row-1}"].font = Font(bold=True, size=12)
    sheet[f"A{start_row-1}"].alignment = Alignment(horizontal="center")

    # Adding rows of data
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            # Adjust column width for readability
            sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Adjust column width for 'Material Name' (second column)
            if c_idx == 2:  # Column B (Material Name)
                sheet.column_dimensions['B'].width = 30  # Increase width of the Material Name column
            else:
                sheet.column_dimensions[chr(64 + c_idx)].width = 15  # Set column width for other columns

            # Align all cells to the center
            sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")

    # Formatting header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FFA500")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border_style = Side(border_style="thin", color="000000")

    # Header formatting
    for cell in sheet[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # Formatting all cells in the table
    for row in sheet.iter_rows(min_row=start_row+1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            # Align all cells to the center
            cell.alignment = Alignment(horizontal="center", vertical="center")

    total_amount = data['Total Price (INR)'].sum()
    return total_amount

# Function to generate and save the Electrical BOQ
def generate_electrical_boq(quantities):
    wb = Workbook()
    # Remove the default sheet created and add a new one for Electrical BOQ
    wb.remove(wb.active)
    ws = wb.create_sheet(title="ElectricalBOQ")
    xlPath=os.path.join(settings.BASE_DIR,'assets','price_of_electrical.xlsx')
    # Load and process the electrical data
    electrical_data = load_data(xlPath, quantities)
    total_electrical = add_boq_data(ws, electrical_data, "ElectricalBOQ")
    
    # Add the Grand Total row
    ws.merge_cells(f"B{ws.max_row + 1}:E{ws.max_row + 1}")
    ws[f"B{ws.max_row}"] = "Grand Total"
    ws[f"B{ws.max_row}"].font = Font(bold=True)
    ws[f"B{ws.max_row}"].alignment = Alignment(horizontal="right")
    ws[f"F{ws.max_row}"] = total_electrical
    ws[f"F{ws.max_row}"].font = Font(bold=True)

    # Save the workbook with no empty sheet
    wb.save("Electrical_BOQ.xlsx")

# %%
def main_electrical(material_result):
    generate_electrical_boq(material_result)
    # Return a success message
    return "Electrical BOQ generated successfully and saved as 'Electrical_BOQ.xlsx'"


