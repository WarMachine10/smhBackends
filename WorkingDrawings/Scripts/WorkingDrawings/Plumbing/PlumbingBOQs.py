
import ezdxf
import ezdxf
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill

def extract_block_counts(dxf_file):
    doc = ezdxf.readfile(dxf_file)
    blocks = {}
    
    for block in doc.modelspace().query('INSERT'):
        block_name = block.dxf.name.strip().upper()
        blocks[block_name] = blocks.get(block_name, 0) + 1
    
    return blocks

def extract_lines_from_layers(dxf_file, layers):
    doc = ezdxf.readfile(dxf_file)
    msp = doc.modelspace()
    lines = {layer: [] for layer in layers}
    
    for entity in msp.query("LINE"):
        if entity.dxf.layer in layers:
            start = (entity.dxf.start.x, entity.dxf.start.y)
            end = (entity.dxf.end.x, entity.dxf.end.y)
            lines[entity.dxf.layer].append((start, end))
    
    return lines

def calculate_line_length(start, end):
    return math.sqrt((end[0] - start[0])**2 + (end[1] - start[1])**2)

def inches_to_feet_and_inches(inches):
    feet = inches // 12
    remaining_inches = inches % 12
    return feet, remaining_inches

def create_layer_length_df(lines, total_floor_height):
    layer_lengths = []
    
    for layer, line_data in lines.items():
        total_length_inches = sum(calculate_line_length(start, end) for start, end in line_data)
        
        if total_length_inches == 0:
            total_length_plus_height = 0
            total_length_str = "0 ft 0 in"
        else:
            feet, inches = inches_to_feet_and_inches(total_length_inches)
            total_length_in_feet = feet + (inches / 12)
            total_length_plus_height = total_length_in_feet + total_floor_height
            total_length_str = f"{feet} ft {inches} in"
        
        layer_lengths.append({
            'Block Name': layer.upper(),
            'Total Length (Feet)': total_length_str,
            'Quantity': round(total_length_plus_height, 2) if total_length_inches != 0 else 0
        })
    
    return pd.DataFrame(layer_lengths)

def main_plumbing_boq( dxf_file,excel_file_path,output_file):
    try:
        # total_floor_height = float(input("Enter the Total Floor Height in feet: "))
        total_floor_height=40
    except ValueError:
        #print("Invalid input! Using default floor height of 10 feet.")
        total_floor_height = 10

    detected_blocks = extract_block_counts(dxf_file)
    layers = {
        "Soil Drain Pipe", "Waste Drain Pipe", "Inlet water Pipe",
        "Outlet water Pipe", "Main Drain Pipe or IC  GT connection",
        "hot water supply", "cold water supply"
    }
    pipe_lengths = create_layer_length_df(extract_lines_from_layers(dxf_file, layers), total_floor_height)
    
    df = pd.read_excel(excel_file_path)
    df.columns = df.columns.str.strip()
    if 'Block Name' not in df.columns:
        raise KeyError("Column 'Block Name' is missing in the Excel file.")
    
    df['Block Name'] = df['Block Name'].str.strip().str.upper()
    df['Quantity'] = 0
    df['Total Price'] = 0
    
    for block_name, count in detected_blocks.items():
        matches = df['Block Name'] == block_name
        if matches.any():
            df.loc[matches & df['Item ID'].notna() & df['Material ID'].isna(), 'Quantity'] = count
            df.loc[matches & df['Item ID'].notna() & df['Material ID'].isna(), 'Total Price'] = df['Unit Price'] * count
    
    for index, row in pipe_lengths.iterrows():
        matches = df['Block Name'] == row['Block Name']
        if matches.any():
            df.loc[matches, 'Quantity'] = row['Quantity']
            df.loc[matches, 'Total Price'] = df['Unit Price'] * row['Quantity']
    
    df.insert(0, 'Sl. No.', range(1, len(df) + 1))
    grand_total = df['Total Price'].sum()
    hardware_cost = grand_total * 0.02
    other_cost = grand_total * 0.02
    
    data_summary = pd.DataFrame([
        {'Sl. No.': '', 'Material Name': 'Hardware (2%)', 'Total Price': hardware_cost},
        {'Sl. No.': '', 'Material Name': 'Other (2%)', 'Total Price': other_cost}
    ])
    grand_total_row = pd.DataFrame([
        {'Sl. No.': '', 'Material Name': 'Grand Total', 'Total Price': grand_total + hardware_cost + other_cost}
    ])
    
    df_final = pd.concat([df, data_summary, grand_total_row], ignore_index=True)
    df_final = df_final.drop(columns=['Material ID', 'Item ID', 'Block Name'], errors='ignore')
    df_final.to_excel(output_file, index=False, startrow=5, engine='openpyxl')
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    img_path = 'SHMLogoB.png'
    try:
        img = Image(img_path)
        img.width, img.height = 500, 50
        img.anchor = 'C2'
        ws.add_image(img)
    except FileNotFoundError:
        print("Warning: Logo image not found.")
    
    ws['B4'] = 'Plumbing BOQ'
    ws['B4'].font = Font(size=14, bold=True)
    ws['B4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B4:F4')
    
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for col_num in range(1, len(df_final.columns) + 1):
        ws.cell(row=6, column=col_num).fill = header_fill
    
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        if row[1].value == 'Grand Total':
            for cell in row:
                cell.font = Font(bold=True)
    
    column_widths = {'A': 5, 'B': 40, 'C': 60, 'D': 15, 'E': 15, 'F': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(output_file)
    #print(f"Plumbing BOQ saved as {output_file}")

# main_plumbing_boq(
#     dxf_file = 'Plumbing_Complete.dxf',
#     excel_file_path = 'BOQs_Plumbing.xlsx',
#     output_file = 'Generated_BOQ_Plumbing.xlsx'
# )